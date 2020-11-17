import openpyxl


def copy_paste():
    wb_read = openpyxl.load_workbook(r"D:\Chrome Downloads\Binu_CurrentHoldings.xlsx")

    sheet_read = wb_read.active
    #print(f'\nsheet active in Fundsindia is  {sheet_read}')

    wb_write = openpyxl.load_workbook(r"H:\Own\Bank.xlsx")
    #sheet_write = wb_write.get_sheet_by_name('Deposit Info- Binu (2)')
    sheet_write = wb_write['Deposit Info- Binu (2)']
    print(f' Sheet active in Bank excel = {sheet_write}')
    #sheet_write = sheet.active
    #print(f'Sheet active in Bank is {sheet_write}')

	#Reading the invested value from the downloaded excel
    value_invested = []
    for i in range(3, 10):
        value_invested.append(sheet_read.cell(row=i, column=5).value)

	#Writing invested values into Bank excel
    x = 0
    for j in range(29, 36):
        write_value = sheet_write.cell(row=j, column=6)
        write_value.value = value_invested[x]
        x = x + 1

	#Reading the actual value from the downloaded excel
    value_return = []
    for i in range(3, 10):
        value_return.append(sheet_read.cell(row=i, column=11).value)

	#Writing actual values into Bank excel
    x = 0
    for j in range(29, 36):
        write_value = sheet_write.cell(row=j, column=7)
        write_value.value = value_return[x]
        x = x + 1

	#Reading the XIRR the downloaded excel
    XIRR = []
    for i in range(3, 10):
        XIRR.append(sheet_read.cell(row=i, column=13).value)
				
	#Writing the XIRR into Bank excel
    x = 0
    for j in range(29, 36):
        write_value = sheet_write.cell(row=j, column=10)
        write_value.value = XIRR[x]/100
        x = x + 1
		
    print('\nBank sheet updated.....\n')
    print(f'Invested value  = {value_invested}')
    print(f'Returns         = {value_return}')
    I = sum(value_invested)
    R = sum(value_return)
    G = R - I
    #print(f'Gain is        = ₹ {round(G, 2)}')
    wb_write.save("H:\Own\Bank.xlsx")
    if R > I:
        print(f"Profit          = ₹{round(G, 2)}",)
    else:
        print(f"Loss            = ₹{round(G, 2)}")
	
	#print(XIRR)


copy_paste()
